"""XAUUSD M1 loading: HistData ``.xlsx`` (EST->UTC) and cached pickles, with a *structural*
in-sample / out-of-sample wall.

Raw source (HistData.com): one ``.xlsx`` per year, sheet named by year, **no header**, six
columns ``datetime, open, high, low, close, volume(=0)``. Timestamps are **fixed EST
(UTC-5, no DST)**; we localize as a fixed offset and convert to UTC. (Ported and re-tested from
the verified V2 loader.) Volume is dropped — HistData volume is identically zero.

The development entry point is :func:`load_is`, which hard-slices the series to ``< oos_start`` so
2023+ rows cannot physically enter development. The locked OOS is reachable only via
:func:`load_oos` / :func:`load_full_m1`.
"""
from __future__ import annotations

import datetime as dt
import glob
import os
import re
import warnings
from typing import List, Sequence

import pandas as pd

from mtf_smc.config import DataConfig

# HistData raw timezone: fixed EST (UTC-5), no daylight saving.
EST_FIXED = dt.timezone(dt.timedelta(hours=-5), name="EST_fixed")
OHLC = ["open", "high", "low", "close"]


# --------------------------------------------------------------------------- #
# Normalization (shared by raw + cache paths)
# --------------------------------------------------------------------------- #
def normalize_ohlc(df: pd.DataFrame, data_tz: str = "UTC") -> pd.DataFrame:
    """Lowercase OHLC, ensure a tz-aware UTC monotonic unique index, validate integrity.

    Raises ``ValueError`` on missing columns or any bar where high/low are inconsistent with
    open/close — we never silently repair data (see ``docs/SPEC.md`` §1.4).
    """
    if not isinstance(df.index, pd.DatetimeIndex):
        raise TypeError("OHLC frame requires a DatetimeIndex (bar start time).")
    out = df.copy()
    out.columns = [str(c).lower() for c in out.columns]
    missing = [c for c in OHLC if c not in out.columns]
    if missing:
        raise ValueError(f"Missing OHLC columns: {missing} (need {OHLC}).")

    idx = out.index
    if idx.tz is None:
        idx = idx.tz_localize(data_tz)
    out.index = idx.tz_convert("UTC")
    out = out[~out.index.duplicated(keep="first")].sort_index()
    for c in OHLC:
        out[c] = out[c].astype("float64")

    bad = (
        (out["high"] < out["low"])
        | (out["high"] < out["open"]) | (out["high"] < out["close"])
        | (out["low"] > out["open"]) | (out["low"] > out["close"])
    )
    if bool(bad.any()):
        raise ValueError(f"OHLC integrity check failed for {int(bad.sum())} bar(s).")
    return out[OHLC]


# --------------------------------------------------------------------------- #
# Raw HistData .xlsx readers (for rebuilding the cache from source)
# --------------------------------------------------------------------------- #
def _xlsx_path(raw_dir: str, symbol: str, year: int) -> str:
    folder = os.path.join(raw_dir, f"HISTDATA_COM_XLSX_{symbol}_M1{year}")
    inner = os.path.join(folder, f"DAT_XLSX_{symbol}_M1_{year}.xlsx")
    if os.path.exists(inner):
        return inner
    cands = glob.glob(os.path.join(folder, "*.xlsx"))
    if not cands:
        raise FileNotFoundError(f"No .xlsx for {symbol} {year} under {folder}")
    return cands[0]


def read_year_xlsx(path: str) -> pd.DataFrame:
    """Read one HistData year ``.xlsx`` -> naive-index DataFrame[open,high,low,close]."""
    import openpyxl

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")  # "no default style" warning is harmless
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        ws.reset_dimensions()  # HistData's dimension tag is often wrong (1x1) -> must reset
        ts: List = []
        o: List = []; h: List = []; lo: List = []; c: List = []
        for row in ws.iter_rows(values_only=True):
            if row[0] is None:
                continue
            ts.append(row[0]); o.append(row[1]); h.append(row[2])
            lo.append(row[3]); c.append(row[4])  # column 6 (volume) ignored
        wb.close()
    return pd.DataFrame({"open": o, "high": h, "low": lo, "close": c},
                        index=pd.DatetimeIndex(ts))


def load_m1_raw(raw_dir, symbol: str, years: Sequence[int], verbose: bool = True) -> pd.DataFrame:
    """Concatenate per-year HistData ``.xlsx`` for ``years``, localize EST->UTC, normalize."""
    frames = []
    for y in years:
        fy = read_year_xlsx(_xlsx_path(str(raw_dir), symbol, y))
        frames.append(fy)
        if verbose:
            print(f"[{symbol}] read {y}: {len(fy)} M1 bars")
    raw = pd.concat(frames, axis=0)
    raw = raw[~raw.index.duplicated(keep="first")].sort_index()
    raw.index = raw.index.tz_localize(EST_FIXED).tz_convert("UTC")  # fixed EST -> UTC
    return normalize_ohlc(raw)


def _discover_years(raw_dir, symbol: str) -> List[int]:
    yrs: List[int] = []
    for p in glob.glob(os.path.join(str(raw_dir), f"HISTDATA_COM_XLSX_{symbol}_M1*")):
        m = re.search(rf"{re.escape(symbol)}_M1(\d{{4}})$", os.path.basename(p))
        if m:
            yrs.append(int(m.group(1)))
    return sorted(set(yrs))


# --------------------------------------------------------------------------- #
# Public loaders
# --------------------------------------------------------------------------- #
def load_full_m1(cfg: DataConfig | None = None, verbose: bool = False) -> pd.DataFrame:
    """Full available M1 series (UTC). Prefers the cache pickle; else builds from raw ``.xlsx``.

    **Do not call this from development code** — it can include OOS rows. Use :func:`load_is`.
    """
    cfg = cfg or DataConfig()
    if cfg.cache_pickle.exists():
        df = normalize_ohlc(pd.read_pickle(cfg.cache_pickle), cfg.data_tz)
        if verbose:
            print(f"[{cfg.symbol}] cache {cfg.cache_pickle.name}: {len(df)} bars "
                  f"{df.index[0]} -> {df.index[-1]}")
        return df
    if cfg.raw_data_dir is None:
        raise FileNotFoundError(
            f"No cache at {cfg.cache_pickle} and raw_data_dir is unset. Set SMC_DATA_DIR or seed "
            f"the cache (see docs/SPEC.md §1.5)."
        )
    years = _discover_years(cfg.raw_data_dir, cfg.symbol)
    if not years:
        raise FileNotFoundError(f"No HistData year folders for {cfg.symbol} under {cfg.raw_data_dir}")
    df = load_m1_raw(cfg.raw_data_dir, cfg.symbol, years, verbose=verbose)
    cfg.data_cache_dir.mkdir(parents=True, exist_ok=True)
    df.to_pickle(cfg.cache_pickle)
    return df


def load_is(cfg: DataConfig | None = None, verbose: bool = False) -> pd.DataFrame:
    """In-sample M1 for ALL development — hard-sliced to ``[is_start, oos_start)``.

    This structural slice guarantees 2023+ (OOS) rows cannot enter development; it is the only
    loader development code should call.
    """
    cfg = cfg or DataConfig()
    df = load_full_m1(cfg, verbose=verbose)
    is_df = df.loc[(df.index >= cfg.is_start) & (df.index < cfg.oos_start)]
    if is_df.empty:
        raise ValueError("In-sample slice is empty — check cache range vs config dates.")
    if verbose:
        print(f"[{cfg.symbol}] IS slice: {len(is_df)} bars {is_df.index[0]} -> {is_df.index[-1]}")
    return is_df


def load_oos(cfg: DataConfig | None = None, verbose: bool = False) -> pd.DataFrame:
    """LOCKED out-of-sample M1 (``[oos_start, oos_end)``). Use ONLY in the final OOS harness."""
    cfg = cfg or DataConfig()
    df = load_full_m1(cfg, verbose=verbose)
    return df.loc[(df.index >= cfg.oos_start) & (df.index < cfg.oos_end)]
